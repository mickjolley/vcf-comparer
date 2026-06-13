# -*- coding: utf-8 -*-
"""
Visual_Phaser.V2.*.py performs comparisons between siblings and cousins and
stores the results in a .xlsx file.

.vcf Files now can be processed.

© 2026 Mick Jolley (mickj1948@gmail.com)

Optimized for speed using a Hybrid Multiprocessing + Multithreading Architecture.
- Multiprocessing: Distributes chromosome analysis across CPU cores.
- Multithreading: Handles concurrent file I/O (DNA loading) and image generation.
"""
import numpy as np
import pandas as pd
import sys
from itertools import combinations
import os
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.utils.cell import column_index_from_string as cs
from openpyxl.utils import get_column_letter as cl
import time
import platform
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
import multiprocessing
import threading
import csv
from openpyxl.drawing.image import Image as XLImage

# Re-import all config variables from the external configuration file
from VCF_Comparer_configV3 import (
    VCF_FILE_PATH, DNA_FILES_PATH, WORKING_DIRECTORY, MAP_PATH, INDIVIDUALS, SUBJECTS,
    CHROMOSOMES, EXCEL_FILE_NAME, SHOW_NO_MATCHES, PC_NO_CALLS_ALLOWED,
    CHROM_TRUE_SIZE, LINEAR_CHROMOSOME, RESOLUTION,
    HIR_CUTOFF, FIR_CUTOFF, PARENTAL_RELATIONSHIP, ROH_CUTOFF,
    FIR_TABLES, FREEZE_COLUMN, LINUX_FONT_STRING,
    HIR_SNP_MIN, FIR_SNP_MIN, MM_DIST
)

# Global cache to store loaded DNA data and a lock to manage concurrent access
worker_dna_cache = {}
cache_lock = threading.Lock()


def get_vcf_individuals(vcf_path):
    """Extracts sample names from the VCF header (#CHROM line)."""
    if not vcf_path or not os.path.exists(vcf_path):
        return []
    try:
        with open(vcf_path, 'r', encoding='utf-8', errors='ignore') as f:
            for line in f:
                if line.startswith('#CHROM'):
                    parts = line.strip().split('\t')
                    if len(parts) > 9:
                        return [p.strip() for p in parts[9:]]
                    break
    except Exception as e:
        print(f"Error reading VCF individuals: {e}")
    return []


def get_subjects_from_path(dna_path):
    """Extracts subject names from filenames in the DNA files path."""
    subjects = []
    if not dna_path or not os.path.isdir(dna_path):
        return []
    try:
        for f in os.listdir(dna_path):
            if '_raw_dna' in f:
                # Assuming format <Name>_raw_dna.txt or similar
                name = f.split('_raw_dna')[0]
                subjects.append(name)
    except Exception as e:
        print(f"Error reading subjects from path: {e}")
    return sorted(list(set(subjects)))


def _looks_like_vcf(file_path):
    if str(file_path).lower().endswith('.vcf'):
        return True

    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as handle:
            for _ in range(20):
                line = handle.readline()
                if not line:
                    break
                stripped = line.strip()
                if stripped.startswith('##fileformat=VCF') or stripped.startswith('#CHROM'):
                    return True
    except OSError:
        return False

    return False

def _parse_vcf_file(file_path, individuals, no_call_val='?'):
    header_columns = None
    separator = '\t'

    with open(file_path, 'r', encoding='utf-8', errors='ignore') as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if not line:
                continue
            if line.startswith('#CHROM'):
                if '\t' in line:
                    header_columns = [c.strip() for c in line.lstrip('#').split('\t')]
                    separator = '\t'
                else:
                    header_columns = [c.strip() for c in line.lstrip('#').split()]
                    separator = r'\s+'
                break

    if not header_columns:
        return

    if separator == '\t':
        raw = pd.read_csv(
            file_path,
            sep='\t',
            comment='#',
            header=None,
            names=header_columns,
            dtype=str,
            low_memory=False,
            keep_default_na=False,
        )
    else:
        raw = pd.read_csv(
            file_path,
            sep=r'\s+',
            comment='#',
            header=None,
            names=header_columns,
            dtype=str,
            low_memory=False,
            keep_default_na=False,
            engine='python',
        )

    if raw.empty:
        return

    chrom_col = _pick_column(raw.columns, ['chrom', 'chromosome'])
    pos_col = _pick_column(raw.columns, ['pos', 'position'])
    id_col = _pick_column(raw.columns, ['id', 'rsid'])
    ref_col = _pick_column(raw.columns, ['ref'])
    alt_col = _pick_column(raw.columns, ['alt'])
    format_col = _pick_column(raw.columns, ['format'])

    if not all([chrom_col, pos_col, ref_col, alt_col]):
        return

    base_df = pd.DataFrame({
        'chromosome': raw[chrom_col].astype(str),
        'position': raw[pos_col].astype(str),
        'rsid': raw[id_col].astype(str) if id_col else '',
    })

    # Default to REF/ALT for non-genotype VCF rows.
    ref_series = raw[ref_col].fillna('').astype(str)
    
    if format_col and len(header_columns) > 9:
        for i in range(9, raw.shape[1]):
            sample_col = header_columns[i]
                        
            if sample_col in individuals:
                format_tokens = raw[format_col].fillna('').astype(str).str.split(':')
                gt_index = format_tokens.apply(lambda toks: toks.index('GT') if 'GT' in toks else -1)
                sample_tokens = raw[sample_col].fillna('').astype(str).str.split(':')
        
                def decode_gt(gt_idx, sample_vals, ref_val, alt_val):
                    choices = [ref_val] + alt_val.split(',')
                    
                    if gt_idx < 0 or gt_idx >= len(sample_vals):
                        return choices[0], choices[1] if len(choices) > 1 else choices[0]
        
                    gt = sample_vals[gt_idx].replace('|', '/').strip()
                    parts = gt.split('/')
        
                    def pick(part):
                        if not part or part == '.':
                            return ''
                        try:
                            idx = int(part)
                            if 0 <= idx < len(choices):
                                return choices[idx]
                        except ValueError:
                            pass
                        return ''
        
                    a1 = pick(parts[0]) if len(parts) > 0 else ''
                    a2 = pick(parts[1]) if len(parts) > 1 else a1
                    
                    # If any allele is missing/invalid, it's a no-call.
                    # _normalize_dna_dataframe will clean these up further.
                    if not a1:
                        a1 = no_call_val
                    if not a2:
                        a2 = a1
                    return a1, a2
        
                decoded = [
                    decode_gt(gt_idx, sample_vals, ref_val, alt_val)
                    for gt_idx, sample_vals, ref_val, alt_val in zip(
                        gt_index.tolist(),
                        sample_tokens.tolist(),
                        ref_series.tolist(),
                        raw[alt_col].fillna('').astype(str).tolist(),
                        strict=True,
                    )
                ]
                
                df = base_df.copy()
                df['allele1'] = [pair[0] for pair in decoded]
                df['allele2'] = [pair[1] for pair in decoded]
                
                # Drop individuals with high no-call percentage
                no_calls = ((df["allele1"] == '?') | (df["allele2"] == '?')).sum()
                total = len(df)
                pc_no_calls = (no_calls / total) * 100 if total > 0 else 0
                if pc_no_calls > PC_NO_CALLS_ALLOWED:
                    print(f"Dropping individual '{sample_col}' due to high no-call percentage: {pc_no_calls:.2f}%")
                    continue
        
                # Populate missing IDs with chromosome-position token.
                missing_id = df['rsid'].isin(['', '.', 'nan', 'None'])
                df.loc[missing_id, 'rsid'] = (
                    df.loc[missing_id, 'chromosome'].astype(str) + ':' + df.loc[missing_id, 'position'].astype(str)
                )
                
                df = _normalize_dna_dataframe(df, no_call_val)
                
                with cache_lock:
                    worker_dna_cache[sample_col] = df.sort_values(by='position').reset_index(drop=True)
                
                print(f"Loaded DNA from VCF for individual: {sample_col}", flush=True)


def _read_raw_dna_table(file_path):
    # VCF files are handled specifically in the main entry point via _parse_vcf_file.
    if _looks_like_vcf(file_path):
        return None

    def parsed_table_looks_usable(df):
        return df is not None and len(df.columns) >= 4

    # Try the common raw-DNA delimiters explicitly, then fall back to auto-detection.
    read_attempts = ['\t', ',']
    for sep in read_attempts:
        try:
            df = pd.read_csv(
                file_path,
                skip_blank_lines=True,
                comment='#',
                header=0,
                low_memory=False,
                dtype=str,
                keep_default_na=False,
                sep=sep,
            )
            if parsed_table_looks_usable(df):
                return df
        except (pd.errors.ParserError, pd.errors.EmptyDataError, UnicodeDecodeError, OSError, ValueError):
            continue

    try:
        df = pd.read_csv(
            file_path,
            skip_blank_lines=True,
            comment='#',
            header=0,
            low_memory=False,
            dtype=str,
            keep_default_na=False,
            sep=None,
            engine='python',
        )
        if parsed_table_looks_usable(df):
            return df
    except (pd.errors.ParserError, pd.errors.EmptyDataError, UnicodeDecodeError, OSError, ValueError):
        pass

    return None

def _pick_column(columns, aliases):
    normalized = {str(col).strip().lower(): col for col in columns}
    for alias in aliases:
        if alias in normalized:
            return normalized[alias]
    return None

def _clean_allele(series, no_call_val='?'):
    no_call_token = str(no_call_val).strip().upper()
    cleaned = series.fillna('').astype(str).str.strip().str.upper()
    cleaned = cleaned.str.replace(r'[^A-Z0-9-]', '', regex=True)

    # Common no-call encodings seen across raw DNA exports.
    no_call_aliases = {'', '-', '--', '0', '00', 'N', 'NN', 'NC', 'NOCALL'}
    cleaned = cleaned.where(~cleaned.isin(no_call_aliases), no_call_token)
    cleaned = cleaned.where(cleaned.isin({'A', 'T', 'C', 'G', no_call_token}), no_call_token)
    return cleaned

def _normalize_dna_dataframe(df, no_call_val='?'):
    # Normalize chromosome labels from multiple vendor formats.
    df['chromosome'] = df['chromosome'].str.strip().str.upper().str.replace('CHR', '', regex=False)
    df['chromosome'] = df['chromosome'].replace({'X': '23', 'XY': '23', 'MT': 'M'})
    df = df[~df['chromosome'].isin(['Y', 'M'])]
    
    # Keep only valid autosomal chromosomes.
    df = df[df['chromosome'].str.isnumeric()]
    df['chromosome'] = df['chromosome'].astype(int)

    # Keep only valid genomic positions.
    df['position'] = pd.to_numeric(df['position'], errors='coerce')
    df = df.dropna(subset=['position'])
    df['position'] = df['position'].astype(int)

    # Clean alleles.
    df['allele1'] = _clean_allele(df['allele1'], no_call_val)
    df['allele2'] = _clean_allele(df['allele2'], no_call_val)
    
    # Drop rows where either allele is a no-call.
    df = df[(df['allele1'] != no_call_val) & (df['allele2'] != no_call_val)]
    
    return df

def agnostic_load_individual_dna(ind, files_path, no_call_val='?', return_error=False):
    """
    Loads and pre-processes DNA for one individual from any supported raw DNA file.
    This parser is delimiter-agnostic (CSV/TAB) and schema-agnostic for common
    consumer DNA exports (Ancestry/23andMe/MyHeritage/FTDNA-like layouts).
    """
    with cache_lock:
        if ind in worker_dna_cache:
            result = (ind, worker_dna_cache[ind])
            if return_error:
                return result + (None,)
            return result

    if not os.path.isdir(files_path):
        if return_error:
            if _looks_like_vcf(files_path):
                return ind, None, f"Individual '{ind}' was not found in the VCF file."
            return ind, None, f"FILES_PATH '{files_path}' is not a directory."
        return ind, None

    file_names = os.listdir(files_path)
    candidates = [name for name in file_names if f"{ind}_raw_dna" in name]
    last_error = f"No matching '*{ind}_raw_dna*' file found in FILES_PATH."
    for filname in candidates:
        this_file = os.path.join(files_path, filname)
        try:
            raw = _read_raw_dna_table(this_file)
            if raw is None or raw.empty:
                last_error = f"{filname}: file could not be parsed or produced no rows."
                continue

            # Resolve columns by common aliases first.
            rsid_col = _pick_column(raw.columns, ['rsid', 'rs#', 'snp'])
            chrom_col = _pick_column(raw.columns, ['chromosome', 'chrom', 'chr'])
            pos_col = _pick_column(raw.columns, ['position', 'pos'])
            allele1_col = _pick_column(raw.columns, ['allele1'])
            allele2_col = _pick_column(raw.columns, ['allele2'])
            genotype_col = _pick_column(raw.columns, ['result', 'genotype', 'alleles', 'allele_pair'])

            # Fallback to column count if headers are non-standard.
            if rsid_col is None or chrom_col is None or pos_col is None:
                cols = list(raw.columns)
                if len(cols) >= 4:
                    rsid_col, chrom_col, pos_col = cols[0], cols[1], cols[2]
                    if len(cols) >= 5:
                        allele1_col, allele2_col = cols[3], cols[4]
                    else:
                        genotype_col = cols[3]
                else:
                    last_error = f"{filname}: missing required columns (need rsid/chromosome/position + alleles or genotype)."
                    continue

            df = pd.DataFrame({
                'rsid': raw[rsid_col].astype(str),
                'chromosome': raw[chrom_col].astype(str),
                'position': raw[pos_col].astype(str),
            })

            if allele1_col is not None and allele2_col is not None:
                df['allele1'] = raw[allele1_col]
                df['allele2'] = raw[allele2_col]
            elif genotype_col is not None:
                genotype = raw[genotype_col].fillna('').astype(str).str.strip().str.upper()
                genotype = genotype.str.replace(r'[^A-Z0-9-]', '', regex=True)
                df['allele1'] = genotype.str[0]
                df['allele2'] = genotype.str[1]
            else:
                last_error = f"{filname}: allele columns were not found and no genotype column was available."
                continue
            
            # Drop individuals with high no-call percentage (consistent with VCF)
            df['allele1'] = _clean_allele(df['allele1'])
            df['allele2'] = _clean_allele(df['allele2'])
            no_calls = ((df["allele1"] == '?') | (df["allele2"] == '?')).sum()
            total = len(df)
            pc_no_calls = (no_calls / total) * 100 if total > 0 else 0
            if pc_no_calls > PC_NO_CALLS_ALLOWED:
                print(f"Dropping individual '{ind}' ({filname}) due to high no-call percentage: {pc_no_calls:.2f}%")
                last_error = f"{filname}: high no-call percentage ({pc_no_calls:.2f}%)."
                continue

            df = _normalize_dna_dataframe(df, no_call_val)

            if df.empty:
                last_error = f"{filname}: no usable autosomal rows after normalization/filtering."
                continue

            print(f"Loaded DNA file successfully: {filname} ({ind})", flush=True)
            result = (ind, df.sort_values(by='position').reset_index(drop=True))
            if return_error:
                return result + (None,)
            return result
        except Exception as e:
            last_error = f"{filname}: {e}"


    if return_error:
        return ind, None, last_error
    return ind, None

def apply_conditions_vectorized(al1x, al2x, al1y, al2y):
    """
    Determines the match type (HIR, FIR, NIR) for a set of alleles using vectorized operations.
    - Crimson: Fully Identical (Both alleles match on both chromosomes).
    - Limegreen: Half Identical (At least one allele matches).
    - Yellow: No match (Different alleles on both chromosomes).
    """
    # cond_nc = (al1x == no_call_val) | (al1y == no_call_val)
    cond_crimson = (al1x == al2x) & (al1y == al2y) & (al1x != al1y)
    cond_limegreen = ((al1x == al1y) & (al2x == al2y)) | ((al1x == al2y) & (al2x == al1y))

    res = np.full(al1x.shape, 'yellow', dtype=object)
    res[cond_limegreen] = 'limegreen'
    res[cond_crimson] = 'crimson'
    # res[cond_nc] = 'limegreen' # Treat no-calls as limegreen for continuity
    return res

def scan_genomes_optimized(dm, chrom, hir_cutoff, fir_cutoff, hir_snp_min, fir_snp_min, mm_dist, dmap_positions, dmap_cms):
    """
    Identifies contiguous segments of matching DNA (HIR and FIR).
    Uses a genetic map (min_map.txt) to calculate distances in centiMorgans (cM).
    - dx: Half-identical regions (HIR)
    - ds: Fully-identical regions (FIR)
    """
    matches = dm["match"].values
    positions = dm["position"].values
    length = len(matches)

    dx, ds = [], []
    nmms = 0
    segflag = fflag = False
    stpos = pos = fstpos = fpos = nsnps = fsnps = mmpos = 0

    def get_dcm(start, end):
        """Interpolates cM distance between two genomic positions."""
        stcm = np.interp(start, dmap_positions, dmap_cms)
        fincm = np.interp(end, dmap_positions, dmap_cms)
        return fincm - stcm

    # Iterative scan through the DNA sequence to find segments
    for i in range(length):
        m, p = matches[i], positions[i]
        if not segflag:
            if m in ('yellow', 'limegreen'):
                nsnps, segflag, stpos = 1, True, p
                if m == 'limegreen':
                    fsnps, fstpos, fflag = 1, p, True
        elif m in ('yellow', 'limegreen'):
            nsnps += 1
            pos = p
            if fflag:
                if m == 'limegreen':
                    fsnps, fpos = fsnps + 1, p
                else:
                    fflag = False
                    if fsnps > fir_snp_min:
                        dcm = get_dcm(fstpos, fpos)
                        if dcm > fir_cutoff:
                            ds.append({"Chr": chrom, "Start Mb": fstpos, "Finish Mb": fpos, "No. SNPs": fsnps, "Length (cM)": round(dcm, 1)})
                    fsnps = 0
            elif m == 'limegreen':
                fsnps, fstpos, fflag = 1, p, True
        else: # m == 'crimson' (No match)
            if fflag:
                if fsnps > fir_snp_min:
                    dcm = get_dcm(fstpos, fpos)
                    if dcm > fir_cutoff:
                        ds.append({"Chr": chrom, "Start Mb": fstpos, "Finish Mb": fpos, "No. SNPs": fsnps, "Length (cM)": round(dcm, 1)})
                fflag, fsnps = False, 0

            nmms += 1
            if nmms == 1:
                mmpos = p
            elif p - mmpos < mm_dist * 1000:
                # End segment if mismatches are too close
                segflag, nmms = False, 0
                if nsnps > hir_snp_min:
                    dcm = get_dcm(stpos, pos)
                    if dcm > hir_cutoff:
                        dx.append({"Chr": chrom, "Start Mb": stpos, "Finish Mb": pos, "No. SNPs": nsnps, "Length (cM)": round(dcm, 1)})
                nsnps = 0
            else:
                nmms, mmpos = 1, p

    # Capture any segments remaining at the end of the chromosome
    if segflag and nsnps > hir_snp_min:
        dcm = get_dcm(stpos, pos)
        if dcm > hir_cutoff:
            dx.append({"Chr": chrom, "Start Mb": stpos, "Finish Mb": pos, "No. SNPs": nsnps, "Length (cM)": round(dcm, 1)})
    if fflag and fsnps > fir_snp_min:
        dcm = get_dcm(fstpos, fpos)
        if dcm > fir_cutoff:
            ds.append({"Chr": chrom, "Start Mb": fstpos, "Finish Mb": fpos, "No. SNPs": fsnps, "Length (cM)": round(dcm, 1)})

    return pd.DataFrame(dx), pd.DataFrame(ds)

def scan_individual_roh(dna_df, chrom, roh_cutoff_mb, dmap_positions, dmap_cms):
    """
    Identifies Runs of Homozygosity (ROH) for a single individual.
    ROHs are contiguous segments where both alleles are identical.
    - roh_cutoff_mb: Minimum segment length in Megabases (Mb).
    Returns (total_cm, list_of_segment_dicts).
    """
    if dna_df.empty:
        return 0.0, []
    
    # Homozygous if allele1 == allele2
    is_homo = (dna_df['allele1'] == dna_df['allele2']).values
    positions = dna_df['position'].values
    
    def get_dcm(start, end):
        stcm = np.interp(start, dmap_positions, dmap_cms)
        fincm = np.interp(end, dmap_positions, dmap_cms)
        return fincm - stcm

    total_roh_cm = 0.0
    roh_segments = []
    in_roh = False
    st_idx = 0
    
    for i in range(len(is_homo)):
        if is_homo[i]:
            if not in_roh:
                st_idx = i
                in_roh = True
        else:
            if in_roh:
                en_idx = i - 1
                st_pos, en_pos = positions[st_idx], positions[en_idx]
                length_mb = (en_pos - st_pos) / 1000000
                if length_mb > roh_cutoff_mb:
                    num_snps = en_idx - st_idx + 1
                    if num_snps > 50 and num_snps/length_mb > 20:
                        dcm = get_dcm(st_pos, en_pos)
                        total_roh_cm += dcm
                        roh_segments.append({
                            "Chr": chrom, "Start Mb": round(st_pos / 1000000, 2), 
                            "Finish Mb": round(en_pos / 1000000, 2), 
                            "No. SNPs": num_snps, "Length (cM)": round(dcm, 1),
                            "Length (Mb)": round(length_mb, 1)
                        })
                in_roh = False
    
    if in_roh:
        en_idx = len(is_homo) - 1
        st_pos, en_pos = positions[st_idx], positions[en_idx]
        length_mb = (en_pos - st_pos) / 1000000
        if length_mb > roh_cutoff_mb:
            num_snps = en_idx - st_idx + 1
            if num_snps > 50 and num_snps/length_mb > 20:
                dcm = get_dcm(st_pos, en_pos)
                total_roh_cm += dcm
                roh_segments.append({
                    "Chr": chrom, "Start Mb": round(st_pos / 1000000, 2), 
                    "Finish Mb": round(en_pos / 1000000, 2), 
                    "No. SNPs": num_snps, "Length (cM)": round(dcm, 1),
                    "Length (Mb)": round(length_mb, 1)
                })
            
    return total_roh_cm, roh_segments

def repair_files_optimized(dm, fir_snp_min, mm_dist):
    """
    Noise reduction: smooths over isolated mismatches or small segments.
    - Fills small gaps in limegreen segments.
    - Reassigns isolated crimson SNPs to yellow.
    """
    matches, positions = dm['match'].values, dm['position'].values
    length = len(matches)
    firs = fir_snp_min // 2
    is_limegreen = (matches == 'limegreen')
    new_matches = matches.copy()

    # Smooth limegreen gaps
    for i in range(firs + 1, length - firs - 1):
        if matches[i] in ('crimson', 'yellow'):
            if np.all(is_limegreen[i-firs : i]) and np.all(is_limegreen[i+1 : i+firs]):
                new_matches[i] = 'limegreen'

    # Identify and reassign isolated crimson SNPs
    crimson_idx = np.where(new_matches == 'crimson')[0]
    if len(crimson_idx) > 0:
        mm_dst = mm_dist * 1000
        for i in range(len(crimson_idx)):
            curr_pos = positions[crimson_idx[i]]
            isolated = True
            if i > 0 and curr_pos - positions[crimson_idx[i-1]] <= mm_dst:
                isolated = False
            if i < len(crimson_idx) - 1 and positions[crimson_idx[i+1]] - curr_pos <= mm_dst:
                isolated = False
            if isolated:
                new_matches[crimson_idx[i]] = 'yellow'

    dm['match'] = new_matches
    return dm

def get_dplot_optimized(q, dtot, dxtot, dstot, pair_name, chrom_true_size, resolution, linear_chromosome, chr_len, siblings):
    """
    Prepares the data for graphical representation (plotting).
    Bins the genetic data into 'pixels' for the output image.
    Determines recombination points by tracking changes in segment types.
    """
    res_val = resolution * 1000
    if chrom_true_size:
        # Use standard chromosome length to ensure all images for this Chr have identical width
        num_bins = int((chr_len / 250000000) * res_val)
    else:
        num_bins = res_val
    num_bins = max(1, num_bins)

    matches, positions = dtot[pair_name].values, dtot['position'].values
    dplot_matches, dplot_positions = np.full(num_bins, 'grey', dtype=object), np.zeros(num_bins)

    # Use fixed number of bins (relative to Chr size if True Size is on) to ensure alignment
    indices = np.linspace(0, len(dtot), num_bins + 1).astype(int)
    for b in range(num_bins):
        start, end = indices[b], indices[b+1]
        if start >= end:
            continue
        
        bin_matches = matches[start:end]
        counts = Counter(bin_matches)
        
        # Prioritize significant matches over 'grey' (missing/no-call)
        if counts['crimson'] > 0:
            dplot_matches[b] = 'crimson'
        elif counts['yellow'] > 0:
            dplot_matches[b] = 'yellow'
        elif counts['limegreen'] > 0:
            dplot_matches[b] = 'limegreen'
        else:
            dplot_matches[b] = 'grey'
        dplot_positions[b] = positions[end-1]

    dplot = pd.DataFrame({'match': dplot_matches, 'position': dplot_positions, 'bar': 'black'})
    for df_tot, color in [(dxtot, 'blue'), (dstot, 'orange')]:
        if len(df_tot) > 0:
            relevant = df_tot[df_tot['pair'] == pair_name]
            for _, row in relevant.iterrows():
                dplot.loc[(dplot['position'] >= row['Start Mb']) & (dplot['position'] <= row['Finish Mb']), 'bar'] = color

    rps, rnames = [], []
    p1, p2 = pair_name.split('-')
    if p1 in siblings and p2 in siblings:
        bar_changes = np.where(dplot['bar'].values[1:] != dplot['bar'].values[:-1])[0] + 1
        for idx in bar_changes:
            rps.append(idx)
            rnames.append(pair_name)

    if linear_chromosome:
        target_res = 10000 if resolution == 10 else 1000
        dplot_final = pd.DataFrame({'match': 'grey', 'bar': 'grey', 'position': np.linspace(0, chr_len, target_res + 1)})
        fracts = (dplot['position'].values / chr_len * target_res).round().astype(int)
        valid = (fracts >= 0) & (fracts <= target_res)
        dplot_final.loc[fracts[valid], 'match'] = dplot['match'].values[valid]
        dplot_final.loc[fracts[valid], 'bar'] = dplot['bar'].values[valid]
        dplot = dplot_final

    return dplot, rps, rnames

def get_roh_dplot(dna_df, roh_segments_df, chrom_true_size, resolution, linear_chromosome, chr_len):
    """
    Prepares data for ROH graphical representation.
    Supports fixed width (normalized) or variable width (true size).
    """
    res_val = resolution * 1000
    if chrom_true_size:
        # Use standard chromosome length to ensure all images for this Chr have identical width
        num_bins = int((chr_len / 250000000) * res_val)
    else:
        num_bins = res_val
    num_bins = max(1, num_bins)

    # Homozygous = limegreen, Heterozygous = crimson
    is_homo = (dna_df['allele1'] == dna_df['allele2']).values
    matches = np.where(is_homo, 'limegreen', 'crimson')
    positions = dna_df['position'].values
    
    dplot_matches, dplot_positions = np.full(num_bins, 'grey', dtype=object), np.zeros(num_bins)

    indices = np.linspace(0, len(dna_df), num_bins + 1).astype(int)
    for b in range(num_bins):
        start, end = indices[b], indices[b+1]
        if start >= end:
            continue
        bin_matches = matches[start:end]
        counts = Counter(bin_matches)
        if counts['crimson'] > 0:
            dplot_matches[b] = 'crimson'
        elif counts['limegreen'] > 0:
            dplot_matches[b] = 'limegreen'
        else:
            dplot_matches[b] = 'grey'
        dplot_positions[b] = positions[end-1]

    dplot = pd.DataFrame({'match': dplot_matches, 'position': dplot_positions, 'bar': 'black'})
    if not roh_segments_df.empty:
        for _, row in roh_segments_df.iterrows():
            st, en = row['Start Mb'] * 1000000, row['Finish Mb'] * 1000000
            dplot.loc[(dplot['position'] >= st) & (dplot['position'] <= en), 'bar'] = 'orange'
            
    if linear_chromosome:
        target_res = 10000 if resolution == 10 else 1000
        dplot_final = pd.DataFrame({'match': 'grey', 'bar': 'grey', 'position': np.linspace(0, chr_len, target_res + 1)})
        fracts = (dplot['position'].values / chr_len * target_res).round().astype(int)
        valid = (fracts >= 0) & (fracts <= target_res)
        dplot_final.loc[fracts[valid], 'match'] = dplot['match'].values[valid]
        dplot_final.loc[fracts[valid], 'bar'] = dplot['bar'].values[valid]
        dplot = dplot_final

    return dplot

def thread_chromosome(chrom, match_pairs, individuals, files_path, map_positions, map_cms, chr_len, siblings, config_params):
    """
    Main worker function for analyzing a single chromosome.
    Orchestrates DNA loading, matching, smoothing, and image preparation.
    Executed in parallel for each chromosome.
    """
    print(f"Analyzing chromosome{chrom}...", flush=True)

    # Step 1: DNA Loading. Uses threading to parallelize disk reads.
    with cache_lock:
        missing_inds = [ind for ind in individuals if ind not in worker_dna_cache]

    if missing_inds:
        with ThreadPoolExecutor(max_workers=min(len(missing_inds), 8)) as threads:
            load_results = threads.map(lambda ind: agnostic_load_individual_dna(ind, files_path), missing_inds)
            with cache_lock:
                for ind, dna_df in load_results:
                    if dna_df is not None:
                        worker_dna_cache[ind] = dna_df

    hir_cutoff = config_params['HIR_CUTOFF']
    fir_cutoff = config_params['FIR_CUTOFF']

    dtot_parts, tables_data = [], []
    dxtot_list, dstot_list = [], []

    # Step 2: Genetic Analysis (CPU-bound)
    # Pre-filter DNA data for the current chromosome to speed up the loop
    current_chrom_dna = {}
    with cache_lock:
        for ind in individuals:
            dna_df = worker_dna_cache.get(ind)
            if dna_df is not None:
                current_chrom_dna[ind] = dna_df[dna_df['chromosome'] == chrom]

    for pair in match_pairs:
        pair_name = f"{pair[0]}-{pair[1]}"
        dna1, dna2 = current_chrom_dna.get(pair[0]), current_chrom_dna.get(pair[1])
        if dna1 is None or dna2 is None:
            continue

        # Merge individual DNA data on common genetic markers
        dm = pd.merge(dna1, dna2, on=("rsid", "chromosome", "position"), suffixes=('_1', '_2'))
        if len(dm) == 0:
            continue

        # Vectorized matching and optional repair
        dm["match"] = apply_conditions_vectorized(dm["allele1_1"].values, dm["allele2_1"].values,
                                                 dm["allele1_2"].values, dm["allele2_2"].values)

        dm = repair_files_optimized(dm, config_params['FIR_SNP_MIN'], config_params['MM_DIST'])

        # Extract HIR and FIR segments
        dx, ds = scan_genomes_optimized(dm, chrom, hir_cutoff, fir_cutoff, config_params['HIR_SNP_MIN'],
                                       config_params['FIR_SNP_MIN'], config_params['MM_DIST'],
                                       map_positions, map_cms)
        
        if dx.empty and ds.empty and not config_params.get('SHOW_NO_MATCHES', False): 
            continue
        
        tables_data.append((pair_name, dx, ds))
        if not dx.empty:
            dx['pair'] = pair_name
        if not ds.empty:
            ds['pair'] = pair_name 
        dxtot_list.append(dx)
        dstot_list.append(ds)
        
        # Ensure 'position' is unique for this pair before using it as an index in pd.concat.
        # Duplicates can occur in low-quality raw DNA files.
        dtot_parts.append(dm[['position', 'match']].drop_duplicates('position').rename(columns={'match': pair_name}))

    if not dtot_parts and not config_params.get('PARENTAL_RELATIONSHIP'):
        return None
    
    dxtot = pd.concat(dxtot_list, ignore_index=True) if dxtot_list else pd.DataFrame()
    dstot = pd.concat(dstot_list, ignore_index=True) if dstot_list else pd.DataFrame()

    # Combine all pair matches into a single chromosome master table
    if dtot_parts:
        # Using pd.concat (outer join) for O(N) performance instead of repeated pd.merge (O(N^2))
        indexed_parts = [p.set_index('position') for p in dtot_parts]
        dtot = pd.concat(indexed_parts, axis=1).reset_index()
        dtot = dtot.fillna('grey').sort_values('position').reset_index(drop=True)
    else:
        dtot = pd.DataFrame(columns=['position'])

    # Step 3: Graphical Preparation and Image Saving. Uses threading for concurrent image saving.
    all_rps, all_rnames, pair_images = [], [], []
    wdir = config_params['WORKING_DIRECTORY'] + "/"
    last_dplot_len = 0
    scale_img_generated = False

    with ThreadPoolExecutor(max_workers=4) as image_threads:
        

        for q, pair in enumerate(match_pairs):
            pair_name = f"{pair[0]}-{pair[1]}"
            if pair_name not in dtot.columns:
                continue

            dplot, rps, rnames = get_dplot_optimized(q, dtot, dxtot, dstot, pair_name, config_params['CHROM_TRUE_SIZE'],
                                                   config_params['RESOLUTION'], config_params['LINEAR_CHROMOSOME'],
                                                   chr_len, siblings)
            all_rps.extend(rps)
            all_rnames.extend(rnames)
            last_dplot_len = len(dplot)
            
            if not scale_img_generated:
                image_threads.submit(get_scale_img, dplot, chrom, wdir)
                scale_img_generated = True
                

            image_threads.submit(get_image_file, dplot, pair_name, chrom, wdir)
            pair_images.append((pair_name, len(dplot)))

        # ROH Analysis (Must be inside image_threads pool for submitting tasks)
        roh_results = {}
        roh_tables = []
        roh_images = []
        if config_params.get('PARENTAL_RELATIONSHIP'):
            roh_cutoff = config_params.get('ROH_CUTOFF', 0)
            show_no_matches = config_params.get('SHOW_NO_MATCHES', False)
            for ind in individuals:
                dna_df = current_chrom_dna.get(ind)
                if dna_df is not None:
                    tcm, segments = scan_individual_roh(dna_df, chrom, roh_cutoff, map_positions, map_cms)
                    roh_results[ind] = tcm
                    if segments or show_no_matches:
                        df_segments = pd.DataFrame(segments) if segments else pd.DataFrame()
                        roh_tables.append((ind, df_segments))
                        
                        # Generate ROH image
                        rdplot = get_roh_dplot(dna_df, df_segments, config_params['CHROM_TRUE_SIZE'], 
                                              config_params['RESOLUTION'], config_params['LINEAR_CHROMOSOME'], chr_len)

                        if not scale_img_generated:
                            image_threads.submit(get_scale_img, rdplot, chrom, wdir)
                            scale_img_generated = True

                        image_threads.submit(get_image_file, rdplot, ind, chrom, wdir)
                        roh_images.append((ind, len(rdplot)))

    return {
        'chrom': chrom, 'tables': tables_data, 'pair_images': pair_images,
        'arp_info': (all_rps, all_rnames, last_dplot_len),
        'dxtot_pairs': list(dxtot['pair'].unique()) if len(dxtot) > 0 else [],
        'dstot_pairs': list(dstot['pair'].unique()) if len(dstot) > 0 else [],
        'roh_results': roh_results,
        'roh_tables': roh_tables,
        'roh_images': roh_images
    }

def get_image_file(dplot, pair_name, chrom, wdir):
    """Generates and saves a visual representation of DNA matches for a sibling pair."""
    img = Image.new("RGB", (len(dplot), 35), color="white")
    draw = ImageDraw.Draw(img)
    colors, bars = dplot['match'].values, dplot['bar'].values
    for i in range(len(dplot)):
        draw.line([(i, 0), (i, 19)], fill=colors[i], width=0)  # SNP match row
        draw.line([(i, 20), (i, 34)], fill=bars[i], width=0)   # Segment row (Blue/Orange)
    img.save(f"{wdir}{pair_name} {chrom}.png")

def get_scale_img(dplot, chrom, wdir):   
    """Generates a genomic scale image showing positions in Megabases (Mb)."""
    img = Image.new("RGB", (len(dplot) + 30, 35), color="white")
    draw = ImageDraw.Draw(img)
    if platform.system() == 'Windows':
        fnt, fnt1 = ImageFont.truetype("arial.ttf", 13), ImageFont.truetype("arial.ttf", 10)
    elif platform.system() == 'Darwin':
        fnt, fnt1 = ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 13), ImageFont.truetype("/System/Library/Fonts/Supplemental/Arial.ttf", 10)
    else:
        fnt, fnt1 = ImageFont.truetype(LINUX_FONT_STRING, 13), ImageFont.truetype(LINUX_FONT_STRING, 10)

    positions = dplot['position'].values
    for i, snp in enumerate(positions):
        if i % 50 == 0:
            draw.text((i, 5), f"{snp / 1000000:0.1f}\n|", font=fnt, fill="black")
        elif i % 5 == 0:
            draw.text((i, 21), '|', font=fnt1, fill="black")
    img.save(f"{wdir}scale {chrom}.png")


def find_next_line(ws, col, addn):
    """Helper to find the next empty row in a given Excel column."""
    lr = 0
    for i in range(ws.max_row, 0, -1):
        if ws.cell(i, col).value is not None:
            lr = i
            break
    return lr + addn

def paste_tables(ws, dx, ds, pair_name, fir_tables, show_no_matches, start_row=1):
    """Pastes segment data (Start, End, SNPs, cM) into the Excel worksheet."""
    side, align = Side(border_style="thin"), Alignment(horizontal="center")
    border = Border(left=side, right=side, top=side, bottom=side)
    if not show_no_matches and len(dx) == 0 and len(ds) == 0:
        return start_row

    current_line = start_row

    def _paste(data, title, line):
        if len(data) == 0:
            return line
        if 'pair' in data.columns:
            data = data.drop('pair', axis=1)
        
        line += 1 # Add spacing
        ws.cell(line, 2).value = title
        for i, col in enumerate(data.columns):
            c = ws.cell(line + 1, 2 + i)
            c.value, c.alignment, c.border = col, align, border
        for i in range(len(data)):
            for j in range(len(data.columns)):
                c = ws.cell(line + 2 + i, 2 + j)
                c.value, c.alignment, c.border = data.iloc[i, j], align, border
        return line + 2 + len(data)

    current_line = _paste(dx, pair_name, current_line)
    if fir_tables and len(ds) > 0:
        current_line = _paste(ds, f"{pair_name} FIR Table", current_line)
    
    return current_line

def paste_image_main(fflag, ws, pair_name, chrom, q, wdir, show_no_matches, dxtot_pairs, dstot_pairs, im_width, dplot_len, start_row=1):
    """Inserts the generated DNA match images into the Excel worksheet."""
    # Column indices depend on whether we have the extra spacer column G
    text_col = 8 if PARENTAL_RELATIONSHIP else 7
    img_col = 9 if PARENTAL_RELATIONSHIP else 8
    text_col_letter = "H" if PARENTAL_RELATIONSHIP else "G"

    if q == 0:
        ws.add_image(XLImage(f"{wdir}scale {chrom}.png"), ws.cell(1, img_col).coordinate)
    if not show_no_matches and pair_name not in dxtot_pairs and pair_name not in dstot_pairs:
        return start_row
    if len(pair_name) > ws.column_dimensions[text_col_letter].width:
        ws.column_dimensions[text_col_letter].width = len(pair_name) + 4

    img = XLImage(f"{wdir}{pair_name} {chrom}.png")
    # Normal sibling placement (offset by 2 rows from previous)
    next_line = max(3, start_row + 2)

    ws.add_image(img, ws.cell(next_line, img_col).coordinate)
    cell = ws.cell(next_line, text_col)
    cell.value, cell.alignment = pair_name, Alignment(horizontal="center")
    
    return next_line

def format_sheet(ws):
    """Sets standard column widths and freezes panes for readability."""
    ws.column_dimensions["A"].width = 1
    if PARENTAL_RELATIONSHIP:
        chars, widths = "BCDEFGH", [5, 11, 12, 11, 13, 14, 14]
    else:
        chars, widths = "BCDEFG", [5, 11, 12, 11, 13, 14]

    for char, w in zip(chars, widths, strict=True):
        ws.column_dimensions[char].width = w

def delete_images(wdir):
    """Clean up: removes temporary .png files generated during the run."""
    for f in os.listdir(wdir):
        if f.endswith(".png"):
            os.remove(os.path.join(wdir, f))

def ensure_visible_worksheet(wb):
    """Guarantee openpyxl can save by keeping at least one visible worksheet."""
    if not wb.worksheets:
        ws = wb.create_sheet("Results")
        ws["A1"] = "No chromosome sheets were generated."
        ws["A2"] = "Check input files and filters in VP_configV1.py."
        return

    visible_sheets = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    if not visible_sheets:
        wb.worksheets[0].sheet_state = "visible"

if __name__ == "__main__":
    start_time = time.time()
    
    # 1. Dynamically load INDIVIDUALS from VCF and SUBJECTS from DNA_FILES_PATH
    vcf_inds = get_vcf_individuals(VCF_FILE_PATH)
    path_subjects = get_subjects_from_path(DNA_FILES_PATH)
    
    # 2. Use configuration values if provided. 
    # Use ['*'] to explicitly load all individuals from VCF.
    if INDIVIDUALS == ['*']:
        current_individuals = vcf_inds
    else:
        current_individuals = INDIVIDUALS if INDIVIDUALS else []
    
    # Use ['*'] to explicitly load all subjects from DNA files path.
    if SUBJECTS == ['*']:
        current_subjects = path_subjects
    else:
        current_subjects = SUBJECTS if SUBJECTS else []
    
    # 3. Define the source and pairs for comparison
    # If both INDIVIDUALS and SUBJECTS are specified, compare each SUBJECT to each INDIVIDUAL.
    if current_individuals and current_subjects:
        if PARENTAL_RELATIONSHIP:
            print(f"\nComputing ROHs for {len(current_subjects)} Subjects from DNA files and {len(current_individuals)} Individuals from VCF...\n", flush=True)
        else:
            print(f"\nComparing {len(current_subjects)} Subjects from DNA files against {len(current_individuals)} Individuals from VCF...\n", flush=True)
        
        # Load VCF individuals
        _parse_vcf_file(VCF_FILE_PATH, current_individuals)
        current_individuals = [i for i in current_individuals if i in worker_dna_cache]
        
        # Load Path subjects
        subject_load_failures = []
        for subj in current_subjects:
            ind, subj_df, error_text = agnostic_load_individual_dna(subj, DNA_FILES_PATH, return_error=True)
            if subj_df is None or subj_df.empty:
                subject_load_failures.append((ind, error_text or "DNA file not found or unreadable."))
            else:
                with cache_lock:
                    worker_dna_cache[ind] = subj_df
        
        if subject_load_failures:
            print("\n[VP_INPUT_ERROR] One or more SUBJECTS could not be loaded.", flush=True)
            for ind, reason in subject_load_failures:
                print(f"[VP_INPUT_ERROR] {ind}: {reason}", flush=True)
            sys.exit(2)
            
        current_subjects = [s for s in current_subjects if s in worker_dna_cache]
        
        if not current_individuals or not current_subjects:
            print("\n[VP_INPUT_ERROR] No individuals or subjects loaded for comparison.", flush=True)
            sys.exit(2)
            
        # Comparison logic: Every Subject vs Every VCF Individual
        match_pairs = []
        if not PARENTAL_RELATIONSHIP:
            for s in current_subjects:
                for i in current_individuals:
                    match_pairs.append((s, i))
        
        # All loaded individuals for processing
        SIBLINGS = current_subjects + current_individuals
        FILES_PATH = DNA_FILES_PATH # Primarily used for fallback, but cache is pre-loaded now

    elif current_individuals:
        # VCF-only mode
        if PARENTAL_RELATIONSHIP:
            print(f"\nComputing ROHs for {len(current_individuals)} Individuals from VCF...\n", flush=True)
        else:
            print(f"\nComparing {len(current_individuals)} Individuals from VCF...\n", flush=True)

        FILES_PATH = VCF_FILE_PATH
        SIBLINGS = current_individuals
        _parse_vcf_file(FILES_PATH, SIBLINGS)
        SIBLINGS = [s for s in SIBLINGS if s in worker_dna_cache]
        
        if not PARENTAL_RELATIONSHIP and len(SIBLINGS) < 2:
            print("\n[VP_INPUT_ERROR] Need at least two valid individuals for comparison.", flush=True)
            sys.exit(2)
        elif not SIBLINGS:
            print("\n[VP_INPUT_ERROR] No valid individuals loaded.", flush=True)
            sys.exit(2)
            
        # If parental relationship, do not compare individuals when subjects list is empty.
        if PARENTAL_RELATIONSHIP and not SUBJECTS:
            match_pairs = []
        else:
            match_pairs = list(combinations(SIBLINGS, 2))
            
    else:
        # Subject-only mode (fallback)
        if PARENTAL_RELATIONSHIP:
            print(f"\nComputing ROHs for {len(current_subjects)} Subjects from DNA files...\n", flush=True)
        else:
            print(f"\nComparing {len(current_subjects)} Subjects from DNA files...\n", flush=True)

        FILES_PATH = DNA_FILES_PATH
        SIBLINGS = current_subjects
        
        # If parental relationship, do not compare subjects when individuals list is empty.
        if PARENTAL_RELATIONSHIP and not INDIVIDUALS:
            match_pairs = []
        else:
            match_pairs = list(combinations(SIBLINGS, 2))

    # Normalize paths
    FILES_PATH, WORKING_DIRECTORY, MAP_PATH = map(os.path.normpath, [FILES_PATH, WORKING_DIRECTORY, MAP_PATH])
    wdir = WORKING_DIRECTORY + "/"
    
    individuals = list(set(SIBLINGS))

    # Pre-flight check: ensure every configured sibling is in cache and usable.
    # (Handling cases where some might have been missed in group loading above)
    sibling_load_failures = []
    for sibling in SIBLINGS:
        # Check cache first
        if sibling in worker_dna_cache:
            continue
            
        ind, sibling_df, error_text = agnostic_load_individual_dna(sibling, FILES_PATH, return_error=True)
        if sibling_df is None or sibling_df.empty:
            sibling_load_failures.append((ind, error_text or "Missing in cache and could not be loaded."))
        else:
            with cache_lock:
                worker_dna_cache[ind] = sibling_df

    if sibling_load_failures:
        print("\n[VP_INPUT_ERROR] One or more SIBLINGS could not be loaded into usable DNA data.", flush=True)
        for ind, reason in sibling_load_failures:
            print(f"[VP_INPUT_ERROR] {ind}: {reason}", flush=True)
        sys.exit(2)

    # Load or create the Excel workbook
    xlname = os.path.join(wdir, f"{EXCEL_FILE_NAME}.xlsx")
    wb = Workbook()
    del wb["Sheet"]

    # Load genetic map (Distance vs genomic position)
    dmap_source = pd.read_csv(os.path.join(MAP_PATH, "min_map.txt"), sep="\t", header=0)
    # Standard chromosome lengths for GRCh37/hg19
    chr_lens = [249250621, 243199373, 198022430, 191154276, 180915260, 171115067, 159138663, 146364022, 141213431, 135534747, 135006516, 133851895, 115169878, 107349540, 102531392, 90354753, 81195210, 78077248, 59128983, 63025520, 48129895, 51304566, 155270560]

    # Setup comparison pairs based on config
    # match_pairs is already defined above in the loading logic

    config_params = {
        'HIR_CUTOFF': HIR_CUTOFF, 'FIR_CUTOFF': FIR_CUTOFF, 
        'HIR_SNP_MIN': HIR_SNP_MIN, 'FIR_SNP_MIN': FIR_SNP_MIN, 'MM_DIST': MM_DIST, 
        'RESOLUTION': RESOLUTION,
        'CHROM_TRUE_SIZE': CHROM_TRUE_SIZE, 'LINEAR_CHROMOSOME': LINEAR_CHROMOSOME, 
        'WORKING_DIRECTORY': WORKING_DIRECTORY,
        'SHOW_NO_MATCHES': SHOW_NO_MATCHES,
        'PARENTAL_RELATIONSHIP': PARENTAL_RELATIONSHIP,
        'ROH_CUTOFF': ROH_CUTOFF
    }

    chrom_list = [int(c) for c in CHROMOSOMES] if CHROMOSOMES else list(range(1, 23))
    print(f"\nProcessing {len(chrom_list)} chromosomes using Threads and Multiprocessing...\nThis will take a few seconds. Please be patient...\n", flush=True)

    # STEP 4: Parallel Processing Loop
    with ThreadPoolExecutor(max_workers=multiprocessing.cpu_count()) as executor:
        futures = {executor.submit(thread_chromosome, c, match_pairs, individuals, FILES_PATH,
                   dmap_source[dmap_source["Chromosome"] == c].sort_values("Position")["Position"].values,
                   dmap_source[dmap_source["Chromosome"] == c].sort_values("Position")["cM"].values,
                   chr_lens[c-1], SIBLINGS, config_params): c for c in chrom_list}

        chromosome_results = {}
        pair_segments = {}
        total_roh_by_ind = {ind: 0.0 for ind in individuals}

        for future in as_completed(futures):
            res = future.result()
            if not res:
                continue
            chromosome_results[res['chrom']] = res
            
            # Aggregate ROH
            if 'roh_results' in res:
                for ind, val in res['roh_results'].items():
                    total_roh_by_ind[ind] += val

        for chrom in sorted(chrom_list):
            res = chromosome_results.get(chrom)
            if not res:
                continue
            
            # Check if there's anything to report (Matches or ROHs)
            has_matches = 'tables' in res and len(res['tables']) > 0
            has_rohs = 'roh_tables' in res and len(res['roh_tables']) > 0
            
            if not has_matches and not has_rohs and not SHOW_NO_MATCHES:
                continue

            print(f"Chromosome {chrom} now merging into Excel...", flush=True)

            # Select or create the worksheet for this chromosome
            ws = wb.create_sheet(f"Chr{chrom}")
            im_width = 0

            ws.freeze_panes = f"{cl(cs(FREEZE_COLUMN)+1)}1"
            format_sheet(ws)

            # Write data tables and images to Excel
            table_row, image_row = 1, 1
            for p_name, dx, ds in res['tables']:
                if p_name not in pair_segments:
                    pair_segments[p_name] = {'hir': [], 'fir': []}
                if not dx.empty:
                    pair_segments[p_name]['hir'].extend(dx.to_dict('records'))
                if not ds.empty:
                    pair_segments[p_name]['fir'].extend(ds.to_dict('records'))
                table_row = paste_tables(ws, dx, ds, p_name, FIR_TABLES, SHOW_NO_MATCHES, start_row=table_row)

            # Paste ROH tables
            if 'roh_tables' in res:
                for ind_name, df_roh in res['roh_tables']:
                    table_row = paste_tables(ws, df_roh, pd.DataFrame(), f"{ind_name} ROH Table", False, True, start_row=table_row)

            fflag = [True] * 24

            # Paste comparison images
            for q, (p_name, dplot_len) in enumerate(res['pair_images']):
                image_row = paste_image_main(fflag, ws, p_name, chrom, q, wdir, SHOW_NO_MATCHES, res['dxtot_pairs'], res['dstot_pairs'], im_width, dplot_len, start_row=image_row)

            # Paste ROH images
            if 'roh_images' in res:
                for q_roh, (roh_name, dplot_len) in enumerate(res['roh_images']):
                    # Offset q by a large number or handle q=0 scale properly
                    # paste_image_main handles scale if q=0. 
                    # If pair_images was empty, first ROH image should handle scale.
                    actual_q = q_roh if not res['pair_images'] else q_roh + 1
                    image_row = paste_image_main(fflag, ws, roh_name, chrom, actual_q, wdir, True, [roh_name], [], im_width, dplot_len, start_row=image_row)

            # Post-processing: Add Recombination Points and Formatting

    # Sort worksheets in numeric chromosome order (Chr1, Chr2, ..., Chr23).
    def _sheet_sort_key(title):
        if title.startswith("Chr") and title[3:].isdigit():
            return (0, int(title[3:]))
        return (1, title)

    sorted_titles = sorted((ws.title for ws in wb.worksheets), key=_sheet_sort_key)
    for idx, title in enumerate(sorted_titles):
        target_sheet = wb[title]
        wb.move_sheet(target_sheet, idx - wb.index(target_sheet))
    
    if PARENTAL_RELATIONSHIP:
        tot_roh = 0    
        for ind in sorted(total_roh_by_ind.keys()):
            tot_roh = tot_roh + total_roh_by_ind[ind]

    # Final Save and Cleanup
    ensure_visible_worksheet(wb)
    if PARENTAL_RELATIONSHIP and tot_roh == 0:
        print('\nThere are no ROHs in any individuals. Excel file not saved.')
    else:
        wb.save(xlname)
    delete_images(wdir)
    total_time = time.time() - start_time
    
    if not PARENTAL_RELATIONSHIP:

        print("\nSummary of Shared DNA:", flush=True)
        summary_data = []
        for pair_name in sorted(pair_segments.keys()):
            segments = pair_segments[pair_name]
            total_shared = 0.0
            # Sum all HIRs
            for h in segments['hir']:
                total_shared += h['Length (cM)']
            
            # Add FIRs only if not contained in an HIR (avoid double counting)
            for f in segments['fir']:
                is_inside = False
                for h in segments['hir']:
                    if h['Chr'] == f['Chr'] and h['Start Mb'] <= f['Start Mb'] and h['Finish Mb'] >= f['Finish Mb']:
                        is_inside = True
                        break
                if not is_inside:
                    total_shared += f['Length (cM)']
            
            print(f"{pair_name}: Shared {total_shared:0.1f} cM", flush=True)
            summary_data.append([pair_name, round(total_shared, 1)])
    
        # Write CSV summary
        csv_path = os.path.join(WORKING_DIRECTORY, f"{EXCEL_FILE_NAME}.csv")
        with open(csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Match", "Total cMs"])
            writer.writerows(summary_data)
    
        print(f"\nData saved to {csv_path}")

    else:
        print("\nSummary of ROH (Runs of Homozygosity):", flush=True)
        roh_summary_data = []
        for ind in sorted(total_roh_by_ind.keys()):
            print(f"{ind}: Total ROH {total_roh_by_ind[ind]:0.1f} cM", flush=True)
            roh_summary_data.append([ind, round(total_roh_by_ind[ind], 1)])
        
        # Write ROH CSV summary
        roh_csv_path = os.path.join(WORKING_DIRECTORY, f"{EXCEL_FILE_NAME}_ROH.csv")
        with open(roh_csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["Individual", "Total ROH cMs"])
            writer.writerows(roh_summary_data)
        print(f"\nROH summary saved to {roh_csv_path}")

    print(f"\nTotal elapsed time = {total_time//60:.0f} min {total_time % 60: .0f} sec.", flush=True)
    print("\nFinished", flush=True)
