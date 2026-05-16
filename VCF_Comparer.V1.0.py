# -*- coding: utf-8 -*-
"""
VCF_Comparer.V1.0.py - Ultra-optimized VCF comparison script.
Identifies Half-Identical Regions (HIR) and Fully-Identical Regions (FIR).
Generates visual plots and an Excel report.

© 2026 Mick Jolley (mickj1948@gmail.com)
Optimized for maximum speed and memory efficiency.
"""

import os
import time
import multiprocessing
from itertools import combinations
import platform
import numpy as np
import pandas as pd
from PIL import Image, ImageDraw, ImageFont
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment
from openpyxl.drawing.image import Image as XLImage
from concurrent.futures import ProcessPoolExecutor, as_completed
from openpyxl.utils.cell import column_index_from_string as cs
from openpyxl.utils import get_column_letter as cl
import csv

# Import all config variables from the external configuration file
from VCF_Comparer_configV1 import (
    DNA_FILES_PATH,
    VCF_FILE_PATH,
    WORKING_DIRECTORY,
    MAP_PATH,
    INDIVIDUALS,
    SUBJECTS,
    CHROMOSOMES,
    EXCEL_FILE_NAME,
    SHOW_NO_MATCHES,
    CHROM_TRUE_SIZE,
    LINEAR_CHROMOSOME,
    RESOLUTION,
    HIR_CUTOFF,
    FIR_CUTOFF,
    X_HIR_CUTOFF,
    X_FIR_CUTOFF,
    SCALE_ON,
    FREEZE_COLUMN,
    HIR_SNP_MIN,
    FIR_SNP_MIN,
    MM_DIST,
    NO_CALL,
)


def _looks_like_vcf(file_path):
    if not os.path.exists(file_path):
        return False
    if str(file_path).lower().endswith(".vcf"):
        return True
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            for _ in range(20):
                line = f.readline()
                if not line:
                    break
                if line.startswith("##fileformat=VCF") or line.startswith("#CHROM"):
                    return True
    except Exception:
        return False
    return False


def _pick_column(columns, aliases):
    normalized = {str(col).strip().lower(): col for col in columns}
    for alias in aliases:
        if alias in normalized:
            return normalized[alias]
    return None


def _clean_allele(series, no_call_val):
    cleaned = series.fillna("").astype(str).str.strip().str.upper()
    cleaned = cleaned.str.replace(r"[^A-Z0-9-]", "", regex=True)
    no_call_aliases = {"", "-", "--", "0", "00", "N", "NN", "NC", "NOCALL"}
    cleaned = cleaned.where(~cleaned.isin(no_call_aliases), no_call_val)
    cleaned = cleaned.where(
        cleaned.isin({"A", "T", "C", "G", no_call_val}), no_call_val
    )
    return cleaned


def _normalize_dna_dataframe(df, no_call_val):
    df["chromosome"] = (
        df["chromosome"]
        .astype(str)
        .str.strip()
        .str.upper()
        .str.replace("CHR", "", regex=False)
    )
    df["chromosome"] = df["chromosome"].replace({"X": "23", "XY": "23", "MT": "M"})
    df = df[~df["chromosome"].isin(["Y", "M"])]
    df = df[df["chromosome"].str.isnumeric()]
    df["chromosome"] = df["chromosome"].astype(int)
    df["position"] = pd.to_numeric(df["position"], errors="coerce")
    df = df.dropna(subset=["position"])
    df["position"] = df["position"].astype(int)
    df["allele1"] = _clean_allele(df["allele1"], no_call_val)
    df["allele2"] = _clean_allele(df["allele2"], no_call_val)
    return df


def _parse_vcf_file(file_path, requested_individuals, no_call_val):
    header_columns = None
    with open(file_path, "r", encoding="utf-8", errors="ignore") as handle:
        for raw_line in handle:
            line = raw_line.strip()
            if line.startswith("#CHROM"):
                header_columns = line.lstrip("#").split("\t")
                break
    if not header_columns:
        return {}
    vcf_samples = header_columns[9:]
    if not requested_individuals or requested_individuals == ["*"]:
        individuals_to_load = vcf_samples
    else:
        individuals_to_load = [i for i in requested_individuals if i in vcf_samples]

    if not individuals_to_load:
        return {}

    print(f"\nLoading {len(individuals_to_load)} individuals from VCF...\n")
    usecols = header_columns[:9] + individuals_to_load
    try:
        raw = pd.read_csv(
            file_path,
            sep="\t",
            comment="#",
            header=None,
            names=header_columns,
            usecols=usecols,
            dtype=str,
            low_memory=False,
            keep_default_na=False,
        )
    except Exception as e:
        print(f"Error reading VCF: {e}")
        return {}

    # Pre-normalize CHROM and POS for the whole VCF once
    raw["chromosome"] = (
        raw["CHROM"]
        .astype(str)
        .str.strip()
        .str.upper()
        .str.replace("CHR", "", regex=False)
    )
    raw["chromosome"] = raw["chromosome"].replace({"X": "23", "XY": "23", "MT": "M"})
    # Filter non-autosomal/X early
    raw = raw[raw["chromosome"].str.isnumeric()].copy()
    raw["chromosome"] = raw["chromosome"].astype(int)
    raw["position"] = pd.to_numeric(raw["POS"], errors="coerce")
    raw = raw.dropna(subset=["position"]).copy()
    raw["position"] = raw["position"].astype(int)

    num_snps = len(raw)
    ref_vals = raw["REF"].values
    alt_list = [alt.split(",") for alt in raw["ALT"].values]
    max_alts = max(len(a) for a in alt_list) if alt_list else 0

    # Create allele matrix: [REF, ALT1, ALT2, ...]
    allele_mat = np.full((num_snps, max_alts + 1), no_call_val, dtype=object)
    allele_mat[:, 0] = ref_vals
    for i, alts in enumerate(alt_list):
        for j, a in enumerate(alts):
            allele_mat[i, j + 1] = a

    # Pre-clean the entire allele matrix once
    for j in range(max_alts + 1):
        allele_mat[:, j] = _clean_allele(
            pd.Series(allele_mat[:, j]), no_call_val
        ).values

    # Find GT index in FORMAT (assuming first row is representative)
    try:
        first_fmt = raw["FORMAT"].iloc[0].split(":")
        gt_idx = first_fmt.index("GT")
    except (ValueError, IndexError):
        print("Error: 'GT' field not found in VCF FORMAT.")
        return {}

    results = {}
    for sample_col in individuals_to_load:
        # Vectorized extraction of GT
        # Split by ':' and take the gt_idx part, then split by '/' or '|'
        gt_col = raw[sample_col].str.split(":").str[gt_idx]
        gt_col = gt_col.str.replace("|", "/", regex=False)
        gt_split = gt_col.str.split("/", n=1, expand=True)

        # Convert indices to integers, handling non-numeric (no-calls)
        g1 = pd.to_numeric(gt_split[0], errors="coerce").fillna(-1).astype(int).values
        g2 = pd.to_numeric(gt_split[1], errors="coerce").fillna(-1).astype(int).values

        # Vectorized lookup in allele_mat
        # Indices for lookup: [row_idx, allele_idx]
        row_indices = np.arange(num_snps)

        def get_alleles(genotypes):
            valid_mask = (genotypes >= 0) & (genotypes <= max_alts)
            res = np.full(num_snps, no_call_val, dtype=object)
            res[valid_mask] = allele_mat[row_indices[valid_mask], genotypes[valid_mask]]
            return res

        a1_vals = get_alleles(g1)
        a2_vals = get_alleles(g2)

        df = pd.DataFrame(
            {
                "chromosome": raw["chromosome"],
                "position": raw["position"],
                "allele1": a1_vals,
                "allele2": a2_vals,
            }
        )
        if df.empty:
            print(f"\nWARNING: VCF Sample '{sample_col}' has no usable autosomal SNPs.")
        else:
            results[sample_col] = df.sort_values(by="position").reset_index(drop=True)
            print(f"Loaded VCF Sample: {sample_col}")
    return results


def _read_raw_dna_table(file_path):
    # VCF files are handled specifically in the main entry point via _parse_vcf_file.
    if _looks_like_vcf(file_path):
        return None

    def parsed_table_looks_usable(df):
        return df is not None and len(df.columns) >= 4

    # Try the common raw-DNA delimiters explicitly, then fall back to auto-detection.
    read_attempts = ["\t", ","]
    for sep in read_attempts:
        try:
            df = pd.read_csv(
                file_path,
                skip_blank_lines=True,
                comment="#",
                header=0,
                low_memory=False,
                dtype=str,
                keep_default_na=False,
                sep=sep,
            )
            if parsed_table_looks_usable(df):
                return df
        except (
            pd.errors.ParserError,
            pd.errors.EmptyDataError,
            UnicodeDecodeError,
            OSError,
            ValueError,
        ):
            continue

    try:
        df = pd.read_csv(
            file_path,
            skip_blank_lines=True,
            comment="#",
            header=0,
            low_memory=False,
            dtype=str,
            keep_default_na=False,
            sep=None,
            engine="python",
        )
        if parsed_table_looks_usable(df):
            return df
    except (
        pd.errors.ParserError,
        pd.errors.EmptyDataError,
        UnicodeDecodeError,
        OSError,
        ValueError,
    ):
        pass

    return None


def agnostic_load_individual_dna(ind, files_path, no_call_val, return_error=False):
    """
    Loads and pre-processes DNA for one individual from any supported raw DNA file.
    This parser is delimiter-agnostic (CSV/TAB) and schema-agnostic for common
    consumer DNA exports (Ancestry/23andMe/MyHeritage/FTDNA-like layouts).
    """

    if not os.path.isdir(files_path):
        err = f"FILES_PATH '{files_path}' is not a directory."
        return (None, err) if return_error else None

    file_names = os.listdir(files_path)
    candidates = [name for name in file_names if f"{ind}_raw_dna" in name]
    last_error = f"No matching '*{ind}_raw_dna*' file found in FILES_PATH."
    for filname in candidates:
        this_file = os.path.join(files_path, filname)
        try:
            raw = _read_raw_dna_table(this_file)
            if raw is None or raw.empty:
                last_error = f"{filname}: file could not be parsed or produced no rows."
                continue

            # Resolve columns by common aliases first.
            rsid_col = _pick_column(raw.columns, ["rsid", "rs#", "snp"])
            chrom_col = _pick_column(raw.columns, ["chromosome", "chrom", "chr"])
            pos_col = _pick_column(raw.columns, ["position", "pos"])
            allele1_col = _pick_column(raw.columns, ["allele1"])
            allele2_col = _pick_column(raw.columns, ["allele2"])
            genotype_col = _pick_column(
                raw.columns, ["result", "genotype", "alleles", "allele_pair"]
            )

            # Fallback to column count if headers are non-standard.
            if rsid_col is None or chrom_col is None or pos_col is None:
                cols = list(raw.columns)
                if len(cols) >= 4:
                    rsid_col, chrom_col, pos_col = cols[0], cols[1], cols[2]
                    if len(cols) >= 5:
                        allele1_col, allele2_col = cols[3], cols[4]
                    else:
                        genotype_col = cols[3]
                else:
                    last_error = f"{filname}: missing required columns (need rsid/chromosome/position + alleles or genotype)."
                    continue

            df = pd.DataFrame(
                {
                    "rsid": raw[rsid_col].astype(str),
                    "chromosome": raw[chrom_col].astype(str),
                    "position": raw[pos_col].astype(str),
                }
            )

            if allele1_col is not None and allele2_col is not None:
                df["allele1"] = raw[allele1_col]
                df["allele2"] = raw[allele2_col]
            elif genotype_col is not None:
                genotype = (
                    raw[genotype_col].fillna("").astype(str).str.strip().str.upper()
                )
                genotype = genotype.str.replace(r"[^A-Z0-9-]", "", regex=True)
                df["allele1"] = genotype.str[0]
                df["allele2"] = genotype.str[1]
            else:
                last_error = f"{filname}: allele columns were not found and no genotype column was available."
                continue

            df = _normalize_dna_dataframe(df, no_call_val)

            if df.empty:
                last_error = f"{filname}: no usable autosomal rows after normalization/filtering."
                continue

            print(f"\nLoaded DNA file successfully: {filname} ({ind})", flush=True)
            result = df.sort_values(by="position").reset_index(drop=True)
            return (result, None) if return_error else result
        except Exception as e:
            last_error = f"{filname}: {e}"

    return (None, last_error) if return_error else None


def find_segments(
    m_vals, p_vals, c_vals, cutoff, snp_min, chrom, mm_dist, is_fir=False
):
    """
    Identifies contiguous matching segments (HIR or FIR) based on SNP data.

    Args:
        m_vals (np.array): Match values (0=mismatch, 1=half-match, 2=full-match/no-call).
        p_vals (np.array): Physical positions in base pairs.
        c_vals (np.array): Genetic positions in centiMorgans (cM).
        cutoff (float): Minimum cM length for a segment to be reported.
        snp_min (int): Minimum number of SNPs required for a segment.
        chrom (int): Chromosome number.
        mm_dist (int): Distance in Kb between mismatches to allow ending a segment.
        is_fir (bool): If True, search for Fully Identical Regions (m_vals == 2).
                       If False, search for Half Identical Regions (m_vals != 0).
    """
    num_snps = len(m_vals)
    if is_fir:
        # FIR detection: Look for contiguous runs where both alleles match (m_vals == 2)
        match_mask = m_vals == 2
        diff = np.diff(match_mask.astype(int), prepend=0, append=0)
        starts = np.where(diff == 1)[0]
        ends = np.where(diff == -1)[0] - 1
        segs = []
        for s, e in zip(starts, ends):
            n = e - s + 1
            if n >= snp_min:
                dcm = c_vals[e] - c_vals[s]
                if dcm >= cutoff:
                    segs.append(
                        {
                            "Chr": chrom,
                            "Start Mb": p_vals[s],
                            "Finish Mb": p_vals[e],
                            "No. SNPs": n,
                            "Length (cM)": round(dcm, 1),
                        }
                    )
        return segs
    else:
        # HIR detection: Look for segments not interrupted by double-mismatches (m_vals == 0)
        mismatch_indices = np.where(m_vals == 0)[0]
        if len(mismatch_indices) == 0:
            # Entire chromosome is a potential match
            if num_snps >= snp_min:
                dcm = c_vals[-1] - c_vals[0]
                if dcm >= cutoff:
                    return [
                        {
                            "Chr": chrom,
                            "Start Mb": p_vals[0],
                            "Finish Mb": p_vals[-1],
                            "No. SNPs": num_snps,
                            "Length (cM)": round(dcm, 1),
                        }
                    ]
            return []

        # Find "stoppers" - mismatches that are close enough to end a segment
        mismatch_positions = p_vals[mismatch_indices]
        dists = np.diff(mismatch_positions)
        close = dists < mm_dist * 1000
        stoppers = mismatch_indices[1:][close]

        # Iterate through segments between stoppers
        segs = []
        last_stop = -1
        for stop in list(stoppers) + [num_snps]:
            s, e = last_stop + 1, stop - 1
            if e >= s:
                n = e - s + 1
                if n >= snp_min:
                    dcm = c_vals[e] - c_vals[s]
                    if dcm >= cutoff:
                        segs.append(
                            {
                                "Chr": chrom,
                                "Start Mb": p_vals[s],
                                "Finish Mb": p_vals[e],
                                "No. SNPs": n,
                                "Length (cM)": round(dcm, 1),
                            }
                        )
            last_stop = stop
        return segs


def analyze_chromosome(
    chrom, match_pairs, map_positions, map_cms, dna_data, config, chr_len
):
    """
    Core analysis loop for a single chromosome. Computes matches for all pairs
    and generates visualization images.
    """
    print(f"Analyzing chromosome {chrom}...")
    hir_cutoff = config["X_HIR_CUTOFF"] if chrom == 23 else config["HIR_CUTOFF"]
    fir_cutoff = config["X_FIR_CUTOFF"] if chrom == 23 else config["FIR_CUTOFF"]

    active_dfs = [df for df in dna_data.values() if not df.empty]
    if not active_dfs:
        return None
    unique_pos = np.unique(np.concatenate([df["position"].values for df in active_dfs]))
    unique_cms = np.interp(unique_pos, map_positions, map_cms)
    pos_to_idx = {p: i for i, p in enumerate(unique_pos)}
    num_pos = len(unique_pos)

    base_map = {config["NO_CALL"]: 0, "A": 1, "C": 2, "G": 3, "T": 4}
    geno_mats = {}
    for ind, df in dna_data.items():
        a1, a2 = np.zeros(num_pos, dtype=np.int8), np.zeros(num_pos, dtype=np.int8)
        indices = [pos_to_idx[p] for p in df["position"].values]
        a1[indices] = [base_map.get(x, 0) for x in df["allele1"].values]
        a2[indices] = [base_map.get(x, 0) for x in df["allele2"].values]
        geno_mats[ind] = (a1, a2)

    results = []
    dt_pairs = []

    # Pre-calculate binning indices for all pairs on this chromosome
    if config["CHROM_TRUE_SIZE"]:
        img_width = int(chr_len / (250000 / config["RESOLUTION"]))
    else:
        img_width = min(config["RESOLUTION"] * 1000, num_pos)

    if config["LINEAR_CHROMOSOME"]:
        # Use physical positions for binning
        bin_edges = np.linspace(0, chr_len, img_width + 1)
        bin_indices = np.digitize(unique_pos, bin_edges) - 1
        bin_indices = np.clip(bin_indices, 0, img_width - 1)
        b_pos = bin_edges[:-1]
    else:
        # Use SNP indices for binning
        bin_indices = np.minimum(
            (np.arange(num_pos) * img_width // num_pos), img_width - 1
        )
        # For SNP-based binning, b_pos is the position of the last SNP in each bin
        b_pos = np.zeros(img_width)
        for i in range(img_width):
            mask = bin_indices == i
            if np.any(mask):
                b_pos[i] = unique_pos[mask][-1]

    COLOR_MAP_RGB = {
        0: [220, 20, 60],  # crimson
        1: [255, 255, 0],  # yellow
        2: [50, 205, 50],  # limegreen
        3: [128, 128, 128],  # grey
    }

    pair_images = []

    for p1, p2 in match_pairs:
        if p1 not in geno_mats or p2 not in geno_mats:
            continue
        a1_1, a2_1 = geno_mats[p1]
        a1_2, a2_2 = geno_mats[p2]

        full = ((a1_1 == a1_2) & (a2_1 == a2_2)) | ((a1_1 == a2_2) & (a2_1 == a1_2))
        mismatch = (
            (a1_1 == a2_1) & (a1_2 == a2_2) & (a1_1 != a1_2) & (a1_1 != 0) & (a1_2 != 0)
        )

        m_vals = np.ones(num_pos, dtype=np.int8)
        m_vals[full] = 2
        m_vals[mismatch] = 0
        nc_mask = (a1_1 == 0) | (a1_2 == 0)
        m_vals[nc_mask] = 2

        p1_pos_arr = dna_data[p1]["position"].values
        p2_pos_arr = dna_data[p2]["position"].values
        # Fast way to find intersection mask
        both_mask = np.isin(unique_pos, p1_pos_arr) & np.isin(unique_pos, p2_pos_arr)

        if not np.any(both_mask):
            continue

        m_sub, p_sub, c_sub = (
            m_vals[both_mask],
            unique_pos[both_mask],
            unique_cms[both_mask],
        )
        dx_df = pd.DataFrame(
            find_segments(
                m_sub,
                p_sub,
                c_sub,
                hir_cutoff,
                config["HIR_SNP_MIN"],
                chrom,
                config["MM_DIST"],
                is_fir=False,
            )
        )
        ds_df = pd.DataFrame(
            find_segments(
                m_sub,
                p_sub,
                c_sub,
                fir_cutoff,
                config["FIR_SNP_MIN"],
                chrom,
                config["MM_DIST"],
                is_fir=True,
            )
        )

        pair_name = f"{p1}-{p2}"
        if not dx_df.empty or not ds_df.empty:
            dt_pairs.append(pair_name)
        results.append((pair_name, dx_df, ds_df))

        # Generate image for this pair
        # match_vals: 0=mismatch, 1=half, 2=full, 3=grey (no data)
        match_vals = np.full(num_pos, 3, dtype=np.int8)
        match_vals[both_mask] = m_vals[both_mask]

        # Binning: take the minimum match value in each bin (prioritize mismatches)
        binned_matches = np.full(img_width, 3, dtype=np.int8)
        # Use a loop over bins for np.min, or a more clever way.
        # np.minimum.reduceat is fast.
        unique_bins, first_indices = np.unique(bin_indices, return_index=True)
        bin_mins = np.minimum.reduceat(match_vals, first_indices)
        binned_matches[unique_bins] = bin_mins

        # bar_vals: 0=black, 1=blue (HIR), 2=orange (FIR), 3=grey (no data)
        bar_vals = np.zeros(img_width, dtype=np.int8)
        if config["LINEAR_CHROMOSOME"]:
            bar_vals[binned_matches == 3] = 3

        for df, val in [(dx_df, 1), (ds_df, 2)]:
            if not df.empty:
                for _, r in df.iterrows():
                    m = (b_pos >= r["Start Mb"]) & (b_pos <= r["Finish Mb"])
                    bar_vals[m] = val

        # Create image array
        img_arr = np.zeros((35, img_width, 3), dtype=np.uint8)
        # Top 20 pixels: match colors
        match_colors = np.array(
            [COLOR_MAP_RGB[v] for v in binned_matches], dtype=np.uint8
        )
        img_arr[:20, :, :] = match_colors[np.newaxis, :, :]

        # Bottom 15 pixels: bar colors
        BAR_COLOR_MAP = {
            0: [0, 0, 0],  # black
            1: [0, 0, 255],  # blue
            2: [255, 165, 0],  # orange
            3: [128, 128, 128],  # grey
        }
        bar_colors = np.array([BAR_COLOR_MAP[v] for v in bar_vals], dtype=np.uint8)
        img_arr[20:, :, :] = bar_colors[np.newaxis, :, :]

        img = Image.fromarray(img_arr)
        img.save(os.path.join(config["WORKING_DIRECTORY"], f"{pair_name} {chrom}.png"))
        pair_images.append((pair_name, img_width))

        # Handle scale image (only for the first pair)
        if pair_name == results[0][0] and config["SCALE_ON"]:
            scale_img = Image.new("RGB", (img_width + 50, 45), color="white")
            s_draw = ImageDraw.Draw(scale_img)
            try:
                if platform.system() == "Windows":
                    fnt = ImageFont.truetype("arial.ttf", 13)
                elif platform.system() == "Darwin":
                    fnt = ImageFont.truetype(
                        "/System/Library/Fonts/Supplemental/Arial.ttf", 13
                    )
                else:
                    fnt = ImageFont.truetype(
                        config.get("LINUX_FONT_STRING", "DejaVuSans.ttf"), 13
                    )
            except Exception:
                fnt = ImageFont.load_default()

            for i in range(0, img_width, 100):
                p = b_pos[i]
                s_draw.text((i, 5), f"{p / 1e6:0.1f}Mb", font=fnt, fill="black")
                s_draw.line([(i, 20), (i, 35)], fill="black", width=1)

            # Add final marker at the end
            if img_width > 0:
                last_idx = img_width - 1
                # If the last marker is too close to a multiple of 100, we might overlap
                if last_idx % 100 > 40 or last_idx < 100:
                    p = (
                        chr_len
                        if (config["LINEAR_CHROMOSOME"] or config["CHROM_TRUE_SIZE"])
                        else b_pos[-1]
                    )
                    s_draw.text(
                        (last_idx, 5), f"{p / 1e6:0.1f}Mb", font=fnt, fill="black"
                    )
                    s_draw.line([(last_idx, 20), (last_idx, 35)], fill="black", width=1)

            scale_img.save(
                os.path.join(config["WORKING_DIRECTORY"], f"scale {chrom}.png")
            )

    if not results:
        return None

    return {
        "chrom": chrom,
        "tables": results,
        "pair_images": pair_images,
        "dt_pairs": dt_pairs,
    }


def main():
    start_time = time.time()
    if not os.path.exists(WORKING_DIRECTORY):
        os.makedirs(WORKING_DIRECTORY)

    # 1. Report missing genetic map
    map_file = os.path.join(MAP_PATH, "min_map.txt")
    if not os.path.exists(map_file):
        print(f"ERROR: Genetic map file not found: {map_file}")

    dmap = (
        pd.read_csv(map_file, sep="\t")
        if os.path.exists(map_file)
        else pd.DataFrame(columns=["Chromosome", "Position", "cM"])
    )

    # 2. Report missing VCF file
    if VCF_FILE_PATH and not os.path.exists(VCF_FILE_PATH):
        print(f"ERROR: VCF file not found: {VCF_FILE_PATH}")

    vcf_to_load = (
        list(set(INDIVIDUALS + SUBJECTS))
        if INDIVIDUALS != ["*"] and SUBJECTS != ["*"]
        else ["*"]
    )
    dna_cache = (
        _parse_vcf_file(VCF_FILE_PATH, vcf_to_load, NO_CALL)
        if _looks_like_vcf(VCF_FILE_PATH)
        else {}
    )

    # 3. Report missing INDIVIDUALS from VCF
    if INDIVIDUALS != ["*"]:
        for i in INDIVIDUALS:
            if i not in dna_cache:
                print(f"\nWARNING: Individual '{i}' not found in VCF sample headers.")

    loaded_vcf_inds = (
        list(dna_cache.keys())
        if INDIVIDUALS == ["*"]
        else [i for i in INDIVIDUALS if i in dna_cache]
    )

    # 3. Report missing DNA directory
    if SUBJECTS and not os.path.isdir(DNA_FILES_PATH):
        print(f"\nERROR: DNA_FILES_PATH is not a directory: {DNA_FILES_PATH}")

    final_subjects = []
    if SUBJECTS == ["*"]:
        if os.path.exists(DNA_FILES_PATH):
            for f in os.listdir(DNA_FILES_PATH):
                name = f.split(".")[0].split("_")[0]
                if name not in dna_cache:
                    df = agnostic_load_individual_dna(name, DNA_FILES_PATH, NO_CALL)
                    if df is not None:
                        dna_cache[name] = df
        final_subjects = [s for s in dna_cache if s not in loaded_vcf_inds]
    else:
        for s in SUBJECTS:
            if s not in dna_cache:
                # Use return_error=True to report why the file couldn't be loaded
                res = agnostic_load_individual_dna(
                    s, DNA_FILES_PATH, NO_CALL, return_error=True
                )
                if res[0] is not None:
                    dna_cache[s] = res[0]
                else:
                    print(f"\nWARNING: Could not load DNA for '{s}': {res[1]}")
            if s in dna_cache:
                final_subjects.append(s)

    if not dna_cache:
        print(
            "\nERROR: No usable DNA data was loaded from any source. Please check your file paths and individual names."
        )
        return

    match_pairs = []
    if final_subjects:
        for s in final_subjects:
            for i in loaded_vcf_inds:
                if s != i:
                    match_pairs.append((s, i))
    else:
        match_pairs = list(combinations(loaded_vcf_inds, 2))

    m_pairs = {}

    for mp in match_pairs:
        pn = mp[0] + "-" + mp[1]
        m_pairs[pn] = 0

    if not match_pairs:
        print("\nNo pairs generated.")
        return
    print("\nPerforming matches. This could take a few minutes. Please be patient...\n")

    config = {
        "HIR_CUTOFF": HIR_CUTOFF,
        "FIR_CUTOFF": FIR_CUTOFF,
        "X_HIR_CUTOFF": X_HIR_CUTOFF,
        "X_FIR_CUTOFF": X_FIR_CUTOFF,
        "HIR_SNP_MIN": HIR_SNP_MIN,
        "FIR_SNP_MIN": FIR_SNP_MIN,
        "MM_DIST": MM_DIST,
        "NO_CALL": NO_CALL,
        "RESOLUTION": RESOLUTION,
        "SCALE_ON": SCALE_ON,
        "WORKING_DIRECTORY": WORKING_DIRECTORY,
        "LINEAR_CHROMOSOME": LINEAR_CHROMOSOME,
        "CHROM_TRUE_SIZE": CHROM_TRUE_SIZE,
    }

    chr_lens = [
        249250621,
        243199373,
        198022430,
        191154276,
        180915260,
        171115067,
        159138663,
        146364022,
        141213431,
        135534747,
        135006516,
        133851895,
        115169878,
        107349540,
        102531392,
        90354753,
        81195210,
        78077248,
        59128983,
        63025520,
        48129895,
        51304566,
        155270560,
    ]

    chrom_list = (
        list(range(1, 24))
        if not CHROMOSOMES or "*" in CHROMOSOMES
        else [int(c) for c in CHROMOSOMES]
    )
    res_map = {}

    with ProcessPoolExecutor(max_workers=multiprocessing.cpu_count()) as ex:
        futures = {}
        for c in chrom_list:
            c_dna = {ind: df[df["chromosome"] == c] for ind, df in dna_cache.items()}
            c_dna = {k: v for k, v in c_dna.items() if not v.empty}
            if not c_dna:
                continue
            m_c = dmap[dmap["Chromosome"] == c].sort_values("Position")
            futures[
                ex.submit(
                    analyze_chromosome,
                    c,
                    match_pairs,
                    m_c["Position"].values,
                    m_c["cM"].values,
                    c_dna,
                    config,
                    chr_lens[c - 1],
                )
            ] = c

        for f in as_completed(futures):
            try:
                r = f.result()
                if r:
                    res_map[r["chrom"]] = r
            except Exception as e:
                print(f"Error Chr {futures[f]}: {e}")

    wb = Workbook()
    del wb["Sheet"]
    side = Side(border_style="thin")
    border = Border(left=side, right=side, top=side, bottom=side)
    align = Alignment(horizontal="center", vertical="center")

    max_pn_len = (
        max([len(f"{p1}-{p2}") for p1, p2 in match_pairs]) if match_pairs else 10
    )
    g_width = max_pn_len + 4

    for c in sorted(res_map.keys()):
        r = res_map[c]
        ws = wb.create_sheet(f"Chr{c}")
        ws.column_dimensions["A"].width = 1
        for char, w in zip("BCDEFG", [5, 11, 12, 11, 13, g_width]):
            ws.column_dimensions[char].width = w
        ws.freeze_panes = f"{cl(cs(FREEZE_COLUMN) + 1)}1"

        curr_row = 1
        for p_name, dx, ds in r["tables"]:
            add_cm = 0
            for data, title in [(dx, p_name), (ds, f"{p_name} FIR")]:
                if data.empty:
                    continue
                ws.cell(curr_row + 1, 2).value = title
                for j, col in enumerate(data.columns):
                    cell = ws.cell(curr_row + 2, 2 + j)
                    cell.value = col
                    cell.alignment = align
                    cell.border = border
                for i in range(len(data)):
                    for j, col in enumerate(data.columns):
                        cell = ws.cell(curr_row + 3 + i, 2 + j)
                        cell.value = data.iloc[i][col]
                        cell.alignment = align
                        cell.border = border
                        if j == 4: 
                            if 'FIR' not in title:
                                # This is an HIR segment, always add to total
                                add_cm += cell.value
                            else:
                                # This is an FIR segment
                                # Check if it is contained within any reported HIR segment (in dx)
                                f_row = data.iloc[i]
                                is_in_hir = False
                                if not dx.empty:
                                    for _, h_row in dx.iterrows():
                                        if (f_row["Start Mb"] >= h_row["Start Mb"] and 
                                            f_row["Finish Mb"] <= h_row["Finish Mb"]):
                                            is_in_hir = True
                                            break
                                if not is_in_hir:
                                    add_cm += cell.value
                curr_row = ws.max_row + 1
            m_pairs[p_name] += round(add_cm, 1)
        nr = 1
        if SCALE_ON and os.path.exists(
            os.path.join(WORKING_DIRECTORY, f"scale {c}.png")
        ):
            ws.add_image(
                XLImage(os.path.join(WORKING_DIRECTORY, f"scale {c}.png")),
                ws.cell(nr, 8).coordinate,
            )
            ws.row_dimensions[nr].height = 30
            nr += 1

        for pn, _ in r["pair_images"]:
            if not SHOW_NO_MATCHES and pn not in r["dt_pairs"]:
                continue
            ip = os.path.join(WORKING_DIRECTORY, f"{pn} {c}.png")
            if os.path.exists(ip):
                ws.add_image(XLImage(ip), ws.cell(nr, 8).coordinate)
                ws.row_dimensions[nr].height = 30
                cell = ws.cell(nr, 7)
                cell.value = pn
                cell.alignment = align
                nr += 1

    xlp = os.path.join(WORKING_DIRECTORY, f"{EXCEL_FILE_NAME}.xlsx")
    wb.save(xlp)

    print()

    for key in m_pairs.keys():
        print(key, f" Total cMs = {m_pairs[key]: .1f}")

    total_time = time.time() - start_time

    print(
        f"\nTotal elapsed time = {total_time // 60:.0f} min {total_time % 60: .0f} sec.",
        flush=True,
    )

    for f in os.listdir(WORKING_DIRECTORY):
        if f.endswith(".png"):
            os.remove(os.path.join(WORKING_DIRECTORY, f))

    with open(
        os.path.join(WORKING_DIRECTORY, f"{EXCEL_FILE_NAME}.csv"), "w", newline=""
    ) as f:
        writer = csv.writer(f)
        writer.writerow(["Match", "Total cMs"])
        for key, val in m_pairs.items():
            writer.writerow([key, val])


if __name__ == "__main__":
    VCF_FILE_PATH = os.path.normpath(VCF_FILE_PATH)
    DNA_FILES_PATH = os.path.normpath(DNA_FILES_PATH)
    WORKING_DIRECTORY = os.path.normpath(WORKING_DIRECTORY)
    MAP_PATH = os.path.normpath(MAP_PATH)

    main()
